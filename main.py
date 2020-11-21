# -*- coding: utf-8 -*-
import re
import jaconv
import inspect
import requests
import os

from datetime import datetime, timedelta
from jsonschema import validate, exceptions
from openpyxl.worksheet.worksheet import Worksheet
from json import dumps

from typing import Dict, List

from util import (SUMMARY_INIT, return_date, get_html_soup, get_file, requests_file, get_weekday, loads_json,
                  dumps_json, month_and_day, jst, print_log, requests_now_data_json, base_url)

# 年代表記の指定
age_display_normal = "代"
age_display_min = "歳未満"
age_display_max = "歳以上"
age_display_unpublished = "非公表"

# Excelファイルのデータの探索を始める最初の行や列の指定
patients_first_row = 2
clusters_first_column = 11
inspections_first_row = 2
main_summary_first_row = 2
columns_name_row = 2

# 医療機関からの発生届が取り下げられたなどの理由により、除外された患者番号を残しておくリスト
exclude_patients = []


class DataManager:
    def __init__(self, patients_sheet: Worksheet, inspections_sheet: Worksheet, summary_sheet: Worksheet):
        # データファイルの設定
        self.patients_sheet = patients_sheet
        self.inspections_sheet = inspections_sheet
        self.summary_sheet = summary_sheet
        # データ量(行数)を調べ始める最初の行の指定
        self.patients_count = patients_first_row
        self.clusters_count = clusters_first_column
        self.inspections_count = inspections_first_row
        self.data_count = main_summary_first_row
        # クラスター一覧を収納するリスト
        self.clusters = []
        # 総病床数 TODO:適宜手動更新が必要なので自動化が望まれる
        self.sickbeds_count = 246
        # 検査数や入院者数などを格納するリスト
        self.summary_values = []
        # 以下、内部変数
        self._patients_json = {}
        self._patients_summary_json = {}
        self._clusters_json = {}
        self._clusters_summary_json = {}
        self._age_json = {}
        self._age_summary_json = {}
        self._inspections_json = {}
        self._inspections_summary_json = {}
        self._main_summary_json = {}
        self._sickbeds_summary_json = {}
        self._current_patients_json = {}
        self._positive_or_negative_json = {}
        self._warning_and_phase_json = {}
        # 初期化(最大行数の取得)
        self.get_patients()
        self.get_clusters()
        self.get_inspections()
        self.get_data_count()

    def json_template_of_patients(self) -> Dict:
        # patients_sheetを用いるデータ向けのテンプレート
        return {
            "data": [],
            "last_update": self.get_patients_last_update()
        }

    def json_template_of_patients_data_dict(self) -> Dict:
        # patients_sheetを用いるデータ向けの、dataがリストではなく辞書型のテンプレート
        return {
            "data": {},
            "last_update": self.get_patients_last_update()
        }

    def json_template_of_inspections(self) -> Dict:
        # patients_sheetを用いるデータ向けのテンプレート
        return {
            "data": [],
            "last_update": self.get_inspections_last_update()
        }

    def dump_and_check_all_data(self) -> bool:
        # xxx_json の名を持つ関数のリストを生成し(_で始まる内部変数は除外する)
        # その後jsonschemaを使ってバリデーションチェックをし、現在のjsonと比較してフラグ(changed_flag)を操作する
        # ちなみに、以降生成するjsonを増やす場合は"_json"で終わる関数と"_"で始まる、関数に対応する内部変数を用意すれば自動で認識される
        json_list = [
            member[0] for member in inspect.getmembers(self) if member[0][-4:] == "json" and member[0][0] != "_"
        ]

        # 変更検知用フラグ
        changed_flag = False

        for json in json_list:
            # 関数は"_json"で終わっているので、それを拡張子に直す必要がある
            json_name = json[:-5] + ".json"
            print_log("data_manager", f"Make {json_name}...")
            # evalで文字列から関数を呼び出している
            made_json = eval("self." + json + "()")

            # 現在デプロイされているjsonを取得し、現在のjsonと比較する
            # 比較結果が「等しくない」のであれば、そのファイルのバリデーションチェックをして出力、
            # 「等しい」のであればそのまま出力する
            now_json = requests_now_data_json(json_name)
            if now_json != made_json:
                changed_flag = True

                # schemaを読み込み、作成したjsonをチェックする。
                print_log("data_manager", f"Validate {json_name}...")
                schema = loads_json(json_name)
                try:
                    validate(made_json, schema)
                except exceptions.ValidationError:
                    raise Exception(f"Check failed {json_name}!")
                print_log("data_manager", f"{json_name} is OK!")
            else:
                print_log("data_manager", f"{json_name} has not changed.")

            # jsonを出力
            print_log("data_manager", f"Dumps {json_name}...")
            dumps_json(json_name, made_json)

        return changed_flag

    # 以下、内部変数を読み取って空ならデータを作成し返す仕組み
    # 直接内部変数を用いるのは、"make_xxx"などでデータを編集するときのみ
    def patients_json(self) -> Dict:
        if not self._patients_json:
            self.make_patients()
        return self._patients_json

    def patients_summary_json(self) -> Dict:
        if not self._patients_summary_json:
            self.make_patients_summary()
        return self._patients_summary_json

    def clusters_json(self) -> Dict:
        if not self._clusters_json:
            self.make_clusters()
        return self._clusters_json

    def clusters_summary_json(self) -> Dict:
        if not self._clusters_summary_json:
            self.make_clusters_summary()
        return self._clusters_summary_json

    def age_json(self) -> Dict:
        if not self._age_json:
            self.make_age()
        return self._age_json

    def age_summary_json(self) -> Dict:
        if not self._age_summary_json:
            self.make_age_summary()
        return self._age_summary_json

    def inspections_json(self) -> Dict:
        if not self._inspections_json:
            self.make_inspections()
        return self._inspections_json

    def inspections_summary_json(self) -> Dict:
        if not self._inspections_summary_json:
            self.make_inspections_summary()
        return self._inspections_summary_json

    def main_summary_json(self) -> Dict:
        if not self._main_summary_json:
            self.make_main_summary()
        return self._main_summary_json

    def sickbeds_summary_json(self) -> Dict:
        if not self._sickbeds_summary_json:
            self.make_sickbeds_summary()
        return self._sickbeds_summary_json

    def current_patients_json(self) -> Dict:
        if not self._current_patients_json:
            self.make_current_patients()
        return self._current_patients_json

    def positive_or_negative_json(self) -> Dict:
        if not self._positive_or_negative_json:
            self.make_positive_or_negative()
        return self._positive_or_negative_json

    def warning_and_phase_json(self) -> Dict:
        if not self._warning_and_phase_json:
            self.make_warning_and_phase()
        return self._warning_and_phase_json

    def make_patients(self) -> None:
        # patients.jsonのデータを作成する
        self._patients_json = self.json_template_of_patients()

        # patients_sheetからデータを読み取っていく
        for i in range(patients_first_row, self.patients_count):
            data = {}
            num = self.patients_sheet.cell(row=i, column=2).value
            # 除外する患者はパスする
            if num in exclude_patients:
                continue
            release_date = return_date(self.patients_sheet.cell(row=i, column=3).value)
            data["No"] = num
            data["リリース日"] = release_date.isoformat()
            data["曜日"] = get_weekday(release_date.weekday())
            # 改行が含まれることがあるので置き換える
            data["居住地"] = str(self.patients_sheet.cell(row=i, column=7).value).replace("\n", "")
            # 年代を一旦取得。「10歳未満」や「90歳以上」、「非公表」と表記されていれば、str型と認識されるので、それを用いて判別する
            age = self.patients_sheet.cell(row=i, column=4).value
            if isinstance(age, int):
                data["年代"] = str(age) + age_display_normal
            else:
                # 「10代未満」の表記を「10歳未満」で統一
                if age[-2:] == age_display_min[1:]:
                    data["年代"] = "10" + age_display_min
                else:
                    # 「90歳以上」と「非公表」はそのまま
                    data["年代"] = age
            data["性別"] = self.patients_sheet.cell(row=i, column=5).value
            data["退院"] = None
            # No.の表記にブレが激しいので、ここで"No."に修正(統一)。また、"・"を"、"に置き換える
            note = self.patients_sheet.cell(row=i, column=11).value
            data["備考"] = None
            if note:
                data["備考"] = re.sub(
                    'NO.|N0.|NO,|N0,|No,', 'No.', str(note)
                ).replace("・", "、").replace("\n", " ")
            data["date"] = release_date.strftime("%Y-%m-%d")

            self._patients_json["data"].append(data)

        # No.1の人からリストに追加していくと、降順になるので、reverseで昇順に戻す
        self._patients_json["data"].reverse()

        # exclude_patientsを入れる
        self._patients_json["exclude_patients"] = sorted(exclude_patients)

    # 以前、データが正常に生成されないことがあったので、inspections_sheetから生成するよう変更済み
    # 念のため、負の遺産として残してある
    # def make_patients_summary(self) -> None:
    #     def make_data(date, value=1):
    #         data = {"日付": date, "小計": value}
    #         return data

    #     self._patients_summary_json = {
    #         "data": [],
    #         "last_update": self.get_last_update()
    #     }

    #     prev_data = {}
    #     for patients_data in sorted(self.patients_json()["data"], key=lambda x: x['date']):
    #         date = patients_data["リリース日"]
    #         if prev_data:
    #             prev_date = datetime.strptime(prev_data["日付"], "%Y-%m-%dT%H:%M:%S+09:00")
    #             patients_zero_days = (datetime.strptime(date, "%Y-%m-%dT%H:%M:%S+09:00") - prev_date).days
    #             if prev_data["日付"] == date:
    #                 prev_data["小計"] += 1
    #                 continue
    #             else:
    #                 self._patients_summary_json["data"].append(prev_data)
    #                 if patients_zero_days >= 2:
    #                     for i in range(1, patients_zero_days):
    #                         self._patients_summary_json["data"].append(
    #                             make_data((prev_date + timedelta(days=i)).replace(tzinfo=jst).isoformat(), 0)
    #                         )
    #         prev_data = make_data(date)
    #     self._patients_summary_json["data"].append(prev_data)
    #     prev_date = datetime.strptime(prev_data["日付"], "%Y-%m-%dT%H:%M:%S+09:00")
    #     patients_zero_days = (datetime.now() - prev_date).days
    #     for i in range(1, patients_zero_days):
    #         self._patients_summary_json["data"].append(
    #             make_data((prev_date + timedelta(days=i)).replace(tzinfo=jst).isoformat(), 0)
    #         )

    def make_patients_summary(self) -> None:
        # patients_summary.jsonの作成
        self._patients_summary_json = self.json_template_of_inspections()

        for inspections_data in self.inspections_json()["data"]:
            date = datetime.strptime(inspections_data["判明日"], "%Y-%m-%d")
            data = {
                "日付": date.replace(tzinfo=jst).isoformat(),
                "小計": inspections_data["陽性確認"]
            }
            self._patients_summary_json["data"].append(data)

    def make_clusters(self) -> None:
        # 内部データテンプレート
        def make_data(date):
            data = {"日付": date}
            # クラスターリストからクラスターを取得し、辞書にはめ込んでいく
            for cluster in self.clusters:
                data[cluster] = 0
            # "None"のものは使われないのでpopで疑似的に削除
            data.pop("None")
            return data

        # clusters.jsonのデータを作成する
        self._clusters_json = self.json_template_of_patients()

        # Excelデータからクラスター一覧に〇がついているところをTrueとし、抜き出す
        patients_cluster_data = []
        for i in range(patients_first_row, self.patients_count):
            # 除外する患者はcontinueで飛ばす
            if self.patients_sheet.cell(row=i, column=2).value in exclude_patients:
                continue
            # 初期化
            cluster_data = {}
            for cluster in self.clusters:
                cluster_data[cluster] = False
            # Noneの削除
            cluster_data.pop("None")
            for j in range(12, self.clusters_count):
                # クラスターの欄に〇があればTrueとする
                if self.patients_sheet.cell(row=i, column=j).value:
                    cluster_data[self.clusters[j - 12]] = True
            # 日時の反映
            cluster_data["date"] = return_date(
                self.patients_sheet.cell(row=i, column=3).value
            ).isoformat()
            patients_cluster_data.append(cluster_data)
        # 患者ごとにデータが保管されるが、順番は関係なく、1日にどこで何人、というデータを抜き出すために、日付順でsortする
        patients_cluster_data.sort(key=lambda x: x['date'])

        # 以前のデータを保管する
        # これは、前の患者データと日付が同じであるか否かを比較するための変数
        prev_data = {}
        for patient in patients_cluster_data:
            date = patient["date"]
            if prev_data:
                prev_date = datetime.strptime(prev_data["日付"], "%Y-%m-%dT%H:%M:%S+09:00")
                # 前のデータと日付が離れている場合、その分0のデータを埋める必要があるので、そのために差を取得する
                patients_zero_days = (datetime.strptime(date, "%Y-%m-%dT%H:%M:%S+09:00") - prev_date).days
                # 前のデータと日付が同じ場合、前のデータに人数を加算していく
                if prev_data["日付"] == date:
                    for j in range(clusters_first_column + 1, self.clusters_count):
                        if self.clusters[j - clusters_first_column - 1] == "None":
                            continue
                        # 以前抜き出したクラスター情報がTrueになっていれば、+1する
                        if patient[self.clusters[j - clusters_first_column - 1]]:
                            prev_data[self.clusters[j - clusters_first_column - 1]] += 1
                    # 加算し終えたら戻る
                    continue
                else:
                    # 前のデータと日付が離れていた場合、前のデータをjsonに登録する
                    self._clusters_json["data"].append(prev_data)
                    # 前のデータとの日付の差が2日以上の場合は空いている日にち分、0を埋める
                    if patients_zero_days >= 2:
                        for i in range(1, patients_zero_days):
                            self._clusters_json["data"].append(
                                make_data((prev_date + timedelta(days=i)).replace(tzinfo=jst).isoformat())
                            )
            # 新しいデータを作成し、前もって前のデータとして格納しておく
            prev_data = make_data(date)
            for j in range(clusters_first_column + 1, self.clusters_count):
                if self.clusters[j - clusters_first_column - 1] == "None":
                    continue
                # 以前抜き出したクラスター情報がTrueになっていれば、+1する
                if patient[self.clusters[j - clusters_first_column - 1]]:
                    prev_data[self.clusters[j - clusters_first_column - 1]] += 1

        # 前のデータをjsonに登録する
        self._clusters_json["data"].append(prev_data)
        prev_date = datetime.strptime(prev_data["日付"], "%Y-%m-%dT%H:%M:%S+09:00")
        # 最終更新のデータから日付が開いている場合、0で埋める
        patients_zero_days = (datetime.now() - prev_date).days - 1
        for i in range(1, patients_zero_days):
            self._clusters_json["data"].append(
                make_data((prev_date + timedelta(days=i)).replace(tzinfo=jst).isoformat())
            )

    def make_clusters_summary(self) -> None:
        # clusters_summary.jsonのデータを作成する
        self._clusters_summary_json = self.json_template_of_patients_data_dict()

        # 初期化
        for cluster in self.clusters:
            self._clusters_summary_json["data"][cluster] = 0
        # clusters.jsonを用いてデータを生成
        for clusters_data in self.clusters_json()["data"]:
            for i in range(len(self.clusters)):
                cluster_name = self.clusters[i]
                if cluster_name == "None":
                    continue
                self._clusters_summary_json["data"][cluster_name] += clusters_data[cluster_name]
        # Noneは削除
        self._clusters_summary_json["data"].pop("None")

    def make_age(self) -> None:
        # age.jsonのデータを作成する
        self._age_json = self.json_template_of_patients_data_dict()

        # 初期化
        for i in range(11):
            if i != 10:
                suffix = age_display_normal
                if i == 0:
                    i = 1
                    suffix = age_display_min
                elif i == 9:
                    suffix = age_display_max
                self._age_json["data"][str(i * 10) + suffix] = 0
            else:
                self._age_json["data"][age_display_unpublished] = 0

        for i in range(patients_first_row, self.patients_count):
            # 除外する患者はcontinueで飛ばす
            if self.patients_sheet.cell(row=i, column=2).value in exclude_patients:
                continue
            age = self.patients_sheet.cell(row=i, column=4).value
            suffix = age_display_normal
            if isinstance(age, str):
                if age == age_display_unpublished:
                    self._age_json["data"][age_display_unpublished] += 1
                    continue
                elif age_display_max in age:
                    age = 90
                    suffix = age_display_max
                else:
                    age = 10
                    suffix = age_display_min
            elif age == 90:
                suffix = age_display_max
            self._age_json["data"][str(age) + suffix] += 1

    def make_age_summary(self) -> None:
        # 内部データテンプレート
        def make_data():
            data = {}
            for i in range(11):
                data[str(i * 10)] = 0
            return data

        # age_summary.jsonを作成する
        self._age_summary_json = {
            "data": {},
            "labels": [],
            "last_update": self.get_patients_last_update()
        }

        # 初期化
        for i in range(11):
            if i != 10:
                suffix = age_display_normal
                if i == 0:
                    i = 1
                    suffix = age_display_min
                elif i == 9:
                    suffix = age_display_max
                self._age_summary_json["data"][str(i * 10) + suffix] = []
            else:
                self._age_summary_json["data"][age_display_unpublished] = []

        # 以前のデータを保管する
        # これは、前の患者データと日付が同じであるか否かを比較するための変数
        patients_age_data = []
        for i in range(patients_first_row, self.patients_count):
            # 10歳未満と、年代非公表者を判別するため、一旦ageに代入し、
            # 年代非公表者は例外として100歳代、10歳未満は便宜上0歳代として扱わせる

            # 除外する患者はcontinueで飛ばす
            if self.patients_sheet.cell(row=i, column=2).value in exclude_patients:
                continue
            age = self.patients_sheet.cell(row=i, column=4).value
            if isinstance(age, str):
                if age == age_display_unpublished:
                    age = 100
                elif age_display_max in age:
                    age = 90
                else:
                    age = 0
            age_data = {
                "年代": age,
                "date": return_date(self.patients_sheet.cell(row=i, column=3).value).isoformat()
            }
            patients_age_data.append(age_data)
        patients_age_data.sort(key=lambda x: x['date'])

        prev_data = {}
        for patient in patients_age_data:
            date = patient["date"]
            if prev_data:
                prev_date = datetime.strptime(prev_data["date"], "%Y-%m-%dT%H:%M:%S+09:00")
                # 前のデータと日付が離れている場合、その分0のデータを埋める必要があるので、そのために差を取得する
                patients_zero_days = (datetime.strptime(date, "%Y-%m-%dT%H:%M:%S+09:00") - prev_date).days
                # 前のデータと日付が同じ場合、前のデータに人数を加算していく
                if prev_data["date"] == date:
                    # 接尾語は扱いづらいので、数字だけに置き換えた辞書で代用している
                    # popで10歳未満や90代以上などの扱いづらいデータを全部0～90に置き換えたものを取り出し、
                    # その後取り出したものに加算し、insertで置き換え直して代入する
                    data = self.pop_age_value()
                    data[str(patient["年代"])] += 1
                    self.insert_age_value(data)
                    continue
                else:
                    # 前のデータとの日付の差が2日以上の場合は空いている日にち分、0を埋める
                    if patients_zero_days >= 2:
                        for i in range(1, patients_zero_days):
                            self.insert_age_value(make_data())
                            self._age_summary_json["labels"].append(
                                month_and_day((prev_date + timedelta(days=i)).replace(tzinfo=jst))
                            )

            data = make_data()
            data[str(patient["年代"])] += 1
            # 作成したデータをリストにinsert
            self.insert_age_value(data)
            # 日時取得のため、前のデータを登録
            prev_data = patient
            self._age_summary_json["labels"].append(
                month_and_day(datetime.strptime(prev_data["date"], "%Y-%m-%dT%H:%M:%S+09:00"))
            )

        prev_date = datetime.strptime(prev_data["date"], "%Y-%m-%dT%H:%M:%S+09:00")
        # 最終更新のデータから日付が開いている場合、0で埋める
        patients_zero_days = (datetime.now() - prev_date).days - 1
        for i in range(1, patients_zero_days):
            self.insert_age_value(make_data())
            self._age_summary_json["labels"].append(
                month_and_day((prev_date + timedelta(days=i)).replace(tzinfo=jst))
            )

    def insert_age_value(self, day_age: Dict) -> None:
        for i in range(11):
            if i != 10:
                j = i
                suffix = age_display_normal
                if i == 0:
                    i = 1
                    suffix = age_display_min
                elif i == 9:
                    suffix = age_display_max
                self._age_summary_json["data"][str(i * 10) + suffix].append(day_age[str(j * 10)])
            else:
                self._age_summary_json["data"][age_display_unpublished].append(day_age[str(i * 10)])

    def pop_age_value(self) -> Dict:
        result = {}
        for i in range(11):
            j = i
            suffix = age_display_normal
            if i == 0:
                i = 1
                suffix = age_display_min
            elif i == 9:
                suffix = age_display_max
            elif i == 10:
                result[str(j * 10)] = self._age_summary_json["data"][age_display_unpublished].pop()
                continue
            result[str(j * 10)] = self._age_summary_json["data"][str(i * 10) + suffix].pop()
        return result

    def make_inspections(self) -> None:
        # inspections.jsonの作成
        self._inspections_json = self.json_template_of_inspections()

        for i in range(inspections_first_row, self.inspections_count):
            date = self.inspections_sheet.cell(row=i, column=1).value
            data = {
                "判明日": date.strftime("%Y-%m-%d"),
                # 0すら入ってない場合はNoneが返ってくるので、0に置き換える
                "地方衛生研究所等": self.inspections_sheet.cell(row=i, column=3).value or 0,
                "民間検査機関等": {
                    "PCR検査": self.inspections_sheet.cell(row=i, column=4).value or 0,
                    "抗原検査": self.inspections_sheet.cell(row=i, column=5).value or 0
                },
                "陽性確認": self.inspections_sheet.cell(row=i, column=6).value or 0
            }
            self._inspections_json["data"].append(data)

    def make_inspections_summary(self) -> None:
        # inspections_summary.jsonの作成
        self._inspections_summary_json = {
            "data": {
                "地方衛生研究所等": [],
                "民間検査機関等": []
            },
            "labels": [],
            "last_update": self.get_inspections_last_update()
        }

        for inspections_data in self.inspections_json()["data"]:
            date = datetime.strptime(inspections_data["判明日"], "%Y-%m-%d")
            self._inspections_summary_json["data"]["地方衛生研究所等"].append(inspections_data["地方衛生研究所等"])
            self._inspections_summary_json["data"]["民間検査機関等"].append(sum(inspections_data["民間検査機関等"].values()))
            self._inspections_summary_json["labels"].append(month_and_day(date))

    def make_main_summary(self) -> None:
        # main_summary.jsonの作成
        # これに関してはテンプレートが大きいのでSUMMARY_INITとして別ファイルに退避している
        self._main_summary_json = SUMMARY_INIT
        self._main_summary_json['last_update'] = self.get_summary_last_update()

        # 以下の式はPDFからデータを取得していた際に使用していたもの。
        # 現在はsummary_sheetの取得に移行しているため、使われていないが、過去にPRしていただいたものとして残している。
        # これ以降、"pdf mode is disabled..."と一緒にコメントアウトされいるものは同意。
        # content = ''.join(self.pdf_texts[3:])
        # self.values = get_numbers_in_text(content)

        # summary_sheetから数値リストを取得
        self.summary_values = self.get_summary_values()
        self.set_summary_values(self._main_summary_json)

    def make_sickbeds_summary(self) -> None:
        # pdf mode is disabled...
        # content = ''.join(self.pdf_texts[3:])
        # self.values = get_numbers_in_text(content)

        # summary_sheetから数値リストを取得
        self.summary_values = self.get_summary_values()
        self._sickbeds_summary_json = {
            "data": {
                "入院患者数": self.summary_values[2],
                "残り病床数": max(0, self.sickbeds_count - self.summary_values[2])
            },
            "last_update": self.get_summary_last_update()
        }

    def make_current_patients(self) -> None:
        # 内部データテンプレート
        def make_data(date, value):
            return {"日付": date, "小計": value}

        # current_patients.jsonのデータを生成する
        self._current_patients_json = self.json_template_of_inspections()

        # まずはinspections_sheetからデータを取得
        for i in range(inspections_first_row, self.inspections_count):
            date = self.inspections_sheet.cell(row=i, column=1).value
            # summary_sheetの最初のデータの日付を超えたらbreak
            summary_date = self.summary_sheet.cell(row=main_summary_first_row, column=1).value
            if date > summary_date:
                break
            if date == summary_date:
                self._current_patients_json["data"].append(
                    make_data(
                        date.replace(tzinfo=jst).isoformat(),
                        self.inspections_sheet.cell(row=i, column=6).value - (
                                self.summary_sheet.cell(row=main_summary_first_row, column=9).value +
                                self.summary_sheet.cell(row=main_summary_first_row, column=10).value
                        )
                    )
                )
            else:
                self._current_patients_json["data"].append(
                    make_data(
                        date.replace(tzinfo=jst).isoformat(),
                        self.inspections_sheet.cell(row=i, column=6).value
                    )
                )

        # 次にsummary_sheetからデータを取得
        for i in range(main_summary_first_row + 1, self.data_count):
            date = self.summary_sheet.cell(row=i, column=1).value
            # 取られるデータが累計値のため、以前の値を引く必要がある
            discharged = (self.summary_sheet.cell(row=i, column=10).value -
                          self.summary_sheet.cell(row=i - 1, column=10).value)
            deaths = (self.summary_sheet.cell(row=i, column=9).value -
                      self.summary_sheet.cell(row=i - 1, column=9).value)
            patients = (self.summary_sheet.cell(row=i, column=4).value -
                        self.summary_sheet.cell(row=i - 1, column=4).value)
            # 退院数と死亡数も引かなければ現在患者数にはならないので、そちらをそれぞれ引く
            # なお、Excel内の「入院患者数」(=現在患者数)は式のため、独自に計算している
            self._current_patients_json["data"].append(
                make_data(date.replace(tzinfo=jst).isoformat(), patients - (discharged + deaths))
            )

    def make_positive_or_negative(self) -> None:
        # positive_or_negative.jsonを生成する
        self._positive_or_negative_json = self.json_template_of_inspections()

        for i in range(inspections_first_row, self.inspections_count):
            date = self.inspections_sheet.cell(row=i, column=1).value.replace(tzinfo=jst)
            data = {"日付": date.isoformat()}
            # それぞれの数値を取得し、Noneの場合は0で置き換える
            official_pcr = self.inspections_sheet.cell(row=i, column=3).value or 0
            unofficial_pcr = self.inspections_sheet.cell(row=i, column=4).value or 0
            unofficial_antigen = self.inspections_sheet.cell(row=i, column=5).value or 0
            positive = self.inspections_sheet.cell(row=i, column=6).value or 0

            negative = official_pcr + unofficial_pcr + unofficial_antigen - positive

            data_len = len(self._positive_or_negative_json["data"])
            data["陽性数"] = positive
            data["陰性数"] = negative

            if data_len < 6:
                positive_average = None
                positive_rate = None
            else:
                week_history = self._positive_or_negative_json["data"][-6:]
                week_history.append(data)
                positive_total = 0
                negative_total = 0
                for day_date in week_history:
                    positive_total += day_date["陽性数"]
                    negative_total += day_date["陰性数"]
                try:
                    positive_average = round(positive_total / 7, 1)
                    positive_rate = round(((positive_total / 7) / ((positive_total + negative_total) / 7)) * 100, 1)
                except ZeroDivisionError:
                    positive_average = 0.0
                    positive_rate = 0.0
            data["7日間平均陽性数"] = positive_average
            data["7日間平均陽性率"] = positive_rate

            self._positive_or_negative_json["data"].append(data)

    def make_warning_and_phase(self) -> None:
        # warning_and_phase.jsonを生成する
        self._warning_and_phase_json = {
            "data": {},
            "last_update": self.get_inspections_last_update()
        }

        warning = 0
        phase = 0

        positive_or_negative_dict = self.positive_or_negative_json()
        average_length = len(positive_or_negative_dict["data"])
        latest_average_patients = positive_or_negative_dict["data"][average_length - 1]["7日間平均陽性数"]

        soup = get_html_soup(base_url)

        real_page_tags = soup.find_all("p", align="center")
        
        for tag in real_page_tags:
            strong = tag.find("strong")
            if strong != None:
                strong_text = strong.get_text()
                if ("感染拡大特別期" in strong_text) and ("【" in strong_text) and ("】" in strong_text):
                    warning = 5
        
        if warning == 5:
            phase = 2
        elif latest_average_patients >= 40:
            warning = 4
            phase = 2
        elif latest_average_patients >= 30:
            warning = 3
            phase = 2
        elif latest_average_patients >= 20:
            warning = 2
            phase = 2
        elif latest_average_patients >= 10:
            warning = 1
            phase = 1
        else:
            warning = 0
            phase = 0

        self._warning_and_phase_json["data"]["警戒基準"] = warning
        self._warning_and_phase_json["data"]["対応の方向性"] = phase

    def get_summary_values(self) -> List:
        values = []
        for i in range(3, 11):
            value = self.summary_sheet.cell(row=self.data_count - 1, column=i).value
            values.append(value)
        return values

    def set_summary_values(self, obj) -> None:
        # リストの先頭の値を"value"にセットする
        obj["value"] = self.summary_values[0]
        # objが辞書型で"children"を持っている場合のみ実行
        if isinstance(obj, dict) and obj.get("children"):
            for child in obj["children"]:
                # 再起させて値をセット
                self.summary_values = self.summary_values[1:]
                self.set_summary_values(child)

    def get_patients_last_update(self) -> str:
        # patients_sheetから"M/D H時現在"の形式で記載されている最終更新日時を取得する
        # クラスターが増えれば端に寄っていき、固定値にすると取得できないので、whileで探索させている
        # また、ファイルによって表記されている行が違うことがあるので、初めに2行目を100列探索させて、見つからなければ次の行を探索させている
        column_num = 16
        data_time_str = ""
        row_num = 1
        while not data_time_str:
            date_time_value = self.patients_sheet.cell(row=row_num, column=column_num).value
            if not date_time_value:
                column_num += 1
                if column_num > 100:
                    column_num = 16
                    row_num += 1
                continue
            # 式(TODAY()-1)が含まれている場合はdatatime型で取得され、時間が別の枠にあるのでそれを取得する
            # TODO: これからこの形が標準になればメインにしたい
            if isinstance(date_time_value, datetime):
                hour_str = ""
                additional_column_num = 1
                while not hour_str:
                    hour_value = self.patients_sheet.cell(row=row_num, column=column_num + additional_column_num).value
                    if not hour_value:
                        additional_column_num += 1
                        continue
                    hour_str = jaconv.z2h(str(hour_value), digit=True, ascii=True)
                    data_time_str = month_and_day(date_time_value) + f" {hour_str}"
                break
            # 数字に全角半角が混じっていることがあるので、半角に統一
            data_time_str = jaconv.z2h(str(date_time_value), digit=True, ascii=True)
        plus_day = 0
        # datetime.strptimeでは24時は読み取れないため。24時を次の日の0時として扱わせる
        if data_time_str[-5:] == "24時現在":
            # 12/31や1/1など、文字数の増減に対応するため、whileで探索させている
            count = 8
            while True:
                try:
                    day_str, hour_str = data_time_str[-count:].split()
                    if day_str.startswith("/"):
                        raise
                    break
                except Exception:
                    count -= 1
            data_time_str = data_time_str[:-count] + day_str + " 0時現在"
            plus_day = 1
        try:
            # 最後に、頭に"2020/"を付け加えてdatetimeに読み取らせている
            # 2021年になった時などどうするかは未定
            # TODO: 年が変わった場合の対応
            last_update = datetime.strptime("2020/" + data_time_str, "%Y/%m/%d %H時現在") + timedelta(days=plus_day)
        except Exception:
            # 兵庫県が日付のフォーマットをミスったままデータをデプロイすることがあるので、それに対する対策
            last_update = datetime.strptime("2020/" + data_time_str, "%Y/%m/%d時現在") + timedelta(days=1)

        return return_date(last_update).isoformat()

    def get_inspections_last_update(self) -> str:
        # 最終データの日の次の日を最終更新日としている
        data_time = self.inspections_sheet.cell(row=self.inspections_count - 1, column=1).value + timedelta(days=1)
        return return_date(data_time).isoformat()

    def get_summary_last_update(self) -> str:
        # pdf mode is disabled...
        # caption = self.pdf_texts[0]
        # dt_vals = get_numbers_in_text(caption)
        # last_update = datetime(datetime.now().year, dt_vals[0], dt_vals[1]) + timedelta(hours=dt_vals[2])
        # return datetime.strftime(last_update, '%Y/%m/%d %H:%M')

        # summary_sheetは一列目が日付、二列目が時間なので、それを読み取って反映させている
        return return_date(
            self.summary_sheet.cell(row=self.data_count - 1, column=1).value +
            timedelta(hours=int(self.summary_sheet.cell(row=self.data_count - 1, column=2).value[:-1]))
        ).isoformat()

    def get_patients(self) -> None:
        # 患者データの行数の取得

        # 患者データの最初の方に空白行がある場合があるので、それを飛ばす。
        global patients_first_row, columns_name_row
        while self.patients_sheet:
            value = self.patients_sheet.cell(row=patients_first_row, column=2).value
            if not value or value == "新型コロナウィルスに感染した患者の状況":
                patients_first_row += 1
            elif value == "番号":
                columns_name_row = patients_first_row
                patients_first_row += 1
            else:
                break
        self.patients_count = patients_first_row

        global exclude_patients
        while self.patients_sheet:
            self.patients_count += 1
            value = self.patients_sheet.cell(row=self.patients_count, column=2).value
            if not value:
                break
            else:
                sub_value_none_count = 0
                for i in range(1, 6):
                    sub_value = self.patients_sheet.cell(row=self.patients_count, column=2+i).value
                    sub_value_none_count += 0 if sub_value else 1
                    # 欠番と書かれるか、5列空白があれば除外患者として登録する
                    if sub_value == "欠番" or sub_value_none_count == 5:
                        exclude_patients.append(value)
                        break

    def get_clusters(self) -> None:
        # クラスターリストの取得とクラスターの列数の取得

        # 一列分、空列があるので、そこを処理するための変数
        none_count = 0
        # "その他/行動歴調査中"がグルーピングされたために、セルが結合されて2行になったので、over_columnとunder_columnの取得が必要になった。
        # under_columnに行きついた際は、under_column一つ目になるので、最初から1を代入しておく
        under_column_count = 1
        while self.patients_sheet:
            self.clusters_count += 1
            over_column = self.patients_sheet.cell(row=columns_name_row, column=self.clusters_count).value
            under_column = self.patients_sheet.cell(row=columns_name_row + 1, column=self.clusters_count).value
            if not over_column:
                # 上のセルが空で、none_countが0の時、読み飛ばす
                if none_count:
                    # 上のセルが空で、none_countが1、更にunder_columnはある時は、under_columnの内容を読み取る
                    if under_column:
                        # under_columnをグルーピングしているover_columnを除外するために、under_columnの数をカウントしている
                        under_column_count += 1
                        self.clusters.append(str(under_column).replace("\n", ""))
                        continue
                    # 上のセルが空で、none_countが1、そのうえunder_columnもない時は、while文を抜ける
                    break
                none_count += 1
            elif over_column == "中核No":
                break

            self.clusters.append(str(over_column).replace("\n", ""))
        # 最後のunder_columnをグルーピングしているover_columnをNoneに置き換える
        self.clusters[-under_column_count] = "None"

    def get_inspections(self) -> None:
        # 検査データの行数の取得
        while self.inspections_sheet:
            self.inspections_count += 1
            value = self.inspections_sheet.cell(row=self.inspections_count, column=1).value
            if not value:
                break

    def get_data_count(self) -> None:
        # サマリーデータの行数の取得
        while self.summary_sheet:
            self.data_count += 1
            value = self.summary_sheet.cell(row=self.data_count, column=1).value
            if not value:
                break


class DataValidator:
    def __init__(self, patients_sheet: Worksheet, inspections_sheet: Worksheet, summary_sheet: Worksheet):
        # データファイルの設定
        self.patients_sheet = patients_sheet
        self.inspections_sheet = inspections_sheet
        self.summary_sheet = summary_sheet
        self.inspections_count = inspections_first_row
        self.slack_webhook = os.environ["SLACK_WEBHOOK"]
        self.get_inspections()

    def check_all_data(self) -> str:
        result_variation = [
            "Found new data warnings!",
            "Some data warnings are solved! But warnings still remains...",
            "Some data warnings are solved! But found new data warnings...",
            "All data warnings are solved!",
            "No new data warnings were found. But warnings still remains...",
            "No new data warnings were found."
        ]

        sheet_list = [
            member[0] for member in inspect.getmembers(self) if member[0][:5] == "check" and member[0][-5:] == "sheet"
        ]
        warnings = []
        result = result_variation[5]
        slack_message = ""
        for sheet in sheet_list:
            print_log("data_validator", f"Run {sheet}...")
            # evalで文字列から関数を呼び出している
            warnings += eval("self." + sheet + "()")

        now_warnings = requests_now_data_json("open_data_warnings.json")
        if not now_warnings:
            now_warnings = []

        fixed_count = 0
        already_fixed_count = 0
        for warning in now_warnings:
            message = warning["message"]
            warnings_message = [w["message"] for w in warnings]
            if message in warnings_message:
                w_index = warnings_message.index(message)
                warnings.pop(w_index)
            elif not warning["fixed"]:
                warning["fixed"] = True
                fixed_count += 1
            else:
                already_fixed_count += 1

        new_warnings = now_warnings + warnings

        dumps_json("open_data_warnings.json", new_warnings)
        if warnings and fixed_count:
            slack_message = f"{fixed_count}個の警告が解決され、新たに{len(warnings)}個の警告が見つかりました。"
            result = result_variation[2]
        elif warnings:
            slack_message = f"新たに{len(warnings)}個の警告が見つかりました。"
            result = result_variation[0]
        elif fixed_count == len(new_warnings) - already_fixed_count:
            slack_message = "すべての警告が解決されました。"
            result = result_variation[3]
        elif fixed_count:
            slack_message = f"{fixed_count}個の警告が解決されましたが、まだいくつかの警告が残っています。"
            result = result_variation[1]
        elif already_fixed_count < len(new_warnings):
            result = result_variation[4]
        elif already_fixed_count == len(new_warnings):
            result = result_variation[5]
        if slack_message:
            if slack_message != "すべての警告が解決されました。":
                slack_message += (
                        "\n" +
                        "詳細は https://stop-covid19-hyogo.github.io/covid19-scraping/open_data_warnings.json をご覧ください。"
                )
            self.slack_notify(slack_message)
        return result

    def check_patients_sheet(self) -> List:
        # データ数がほかのデータと相違ないか、データ形式が間違っていないか
        patients_warning = []
        patients_column = patients_first_row
        count = 1
        patients_count = 0
        prev_date = None

        def add_warning_message(message: str, option_file: str = ""):
            patients_warning.append(
                {
                    "message": message,
                    "file": "patients" + (f", {option_file}" if option_file else ""),
                    "fixed": False
                }
            )

        # 全体として、データ数の確認数をする
        while True:
            num = self.patients_sheet.cell(row=patients_column, column=2).value
            prev_num = self.patients_sheet.cell(row=patients_column - 1, column=2).value
            if isinstance(prev_num, int) and isinstance(num, int):
                if prev_num - 1 != num:
                    add_warning_message(
                        f"{num}番の患者番号に間違いがある可能性があります。" +
                        f"上の行の番号との差が1ではありません(上の行の番号:{prev_num})"
                    )
            if num in exclude_patients:
                patients_column += 1
                continue
            if num is not None:
                date = return_date(self.patients_sheet.cell(row=patients_column, column=3).value)
                # ここで、データ単体の確認をする
                # 居住地がおかしくないか
                residence = self.patients_sheet.cell(row=patients_column, column=7).value
                if not residence.endswith(("市", "町", "都", "道", "府", "県", "市外", "県外", "健康福祉事務所管内")):
                    if residence not in ["調査中", "非公表"]:
                        add_warning_message(
                            f"{num}番の患者データに間違いがある可能性があります。" +
                            f"居住地が定型に当てはまっていません({residence})"
                        )
                # 性別はおかしくないか
                sex = self.patients_sheet.cell(row=patients_column, column=5).value
                if sex not in ["男性", "女性", "非公表"]:
                    add_warning_message(
                        f"{num}番の患者データに間違いがある可能性があります。" +
                        f"性別が不適切です({sex})"
                    )
                # 年代はおかしくないか
                age = self.patients_sheet.cell(row=patients_column, column=4).value
                if isinstance(age, str):
                    if age in [age_display_unpublished, f"{10}{age_display_min}", f"{90}{age_display_max}"]:
                        pass
                    else:
                        add_warning_message(
                            f"{num}番の患者データに間違いがある可能性があります。" +
                            f"年代が不適切です({age})"
                        )
                elif isinstance(age, int):
                    pass
                else:
                    add_warning_message(
                        f"{num}番の患者データに間違いがある可能性があります。" +
                        f"年代が不適切です({age})"
                    )
                # 管轄はおかしくないか
                jurisdiction = str(self.patients_sheet.cell(row=patients_column, column=6).value)
                # 現状判明している管轄(どうやら健康福祉事務所だけではないらしい)
                # 新たなものは分かり次第追加する
                if jurisdiction not in [
                    "芦屋", "宝塚", "伊丹", "加古川", "加東", "中播磨", "龍野", "赤穂",
                    "豊岡", "朝来", "丹波", "洲本", "神戸", "姫路", "尼崎", "西宮", "明石"
                ]:
                    add_warning_message(
                        f"{num}番の患者データに間違いがある可能性があります。" +
                        f"管轄が不適切です({jurisdiction})"
                    )
                # 発症日はおかしくないか
                onset_date = self.patients_sheet.cell(row=patients_column, column=9).value
                if return_date(onset_date) is None:
                    if onset_date not in ["症状なし", "調査中", "非公表"]:
                        add_warning_message(
                            f"{num}番の患者データに間違いがある可能性があります。" +
                            f"発症日が不適切です({onset_date})"
                        )
                # 渡航歴はおかしくないか
                travel_history = str(self.patients_sheet.cell(row=patients_column, column=10).value)
                if travel_history not in ["あり", "なし", "調査中", "不明", "非公表"]:
                    add_warning_message(
                        f"{num}番の患者データに間違いがある可能性があります。" +
                        f"渡航歴が不適切です({travel_history})"
                    )
            else:
                date = None
            inspections_last_date = return_date(
                self.inspections_sheet.cell(row=self.inspections_count - 1, column=1).value
            )
            if prev_date is None or prev_date == date:
                patients_count += 1
            elif inspections_last_date < prev_date:
                # inspections_sheetの公開がpatients_sheetの公開より遅い場合は、同じ日のデータが見つけられないので検証をパスする
                # なお、patients_countを1にするのはprev_dateがinspections_last_dateと同じ日になった時のため(elseの時と同じ処理)
                patients_count = 1
            else:
                # 感染者0の日もあるので、感染者があった日のデータに合うようにする
                while prev_date != return_date(
                        self.inspections_sheet.cell(row=self.inspections_count - count, column=1).value
                ):
                    count += 1

                patients_count_from_inspections_sheet = self.inspections_sheet.cell(
                    row=self.inspections_count - count, column=6
                ).value
                if patients_count != patients_count_from_inspections_sheet:
                    add_warning_message(
                        f"患者データの{month_and_day(prev_date)}の分に間違いがある可能性があります。" +
                        f"小計が合いません(差分:{patients_count_from_inspections_sheet - patients_count})",
                        "inspections"
                    )
                if date is None:
                    break
                patients_count = 1
            prev_date = date
            patients_column += 1

        patients_warning.reverse()
        return patients_warning

    def check_inspections_sheet(self) -> List:
        inspections_warning = []
        inspections_row = inspections_first_row
        inspections_total = 0
        patients_total = 0
        count = 0

        patients_warnings = [m["message"] for m in self.check_patients_sheet()]

        def add_warning_message(message: str, option_file: str = "", only_option_file: bool = False):
            inspections_warning.append(
                {
                    "message": message,
                    "file": ("inspections" if not only_option_file else "") +
                            (f"{', ' if not only_option_file else ''}{option_file}" if option_file else ""),
                    "fixed": False
                }
            )

        while True:
            date = return_date(self.inspections_sheet.cell(row=inspections_row, column=1).value)
            summary_date = return_date(self.summary_sheet.cell(row=main_summary_first_row + count, column=1).value)
            if summary_date is None or date is None:
                break

            # データの取得
            inspections_subtotal = self.inspections_sheet.cell(row=inspections_row, column=2).value or 0
            official_pcr = self.inspections_sheet.cell(row=inspections_row, column=3).value or 0
            unofficial_pcr = self.inspections_sheet.cell(row=inspections_row, column=4).value or 0
            unofficial_antigen = self.inspections_sheet.cell(row=inspections_row, column=5).value or 0
            patients_in_day = self.inspections_sheet.cell(row=inspections_row, column=6).value or 0
            subtotal = official_pcr + unofficial_pcr + unofficial_antigen

            if inspections_subtotal != subtotal:
                add_warning_message(
                    f"{month_and_day(date)}の検査数に間違いがある可能性があります。" +
                    f"小計(1日ごとの合計)が合いません(差分:{inspections_subtotal - subtotal})"
                )

            inspections_total += inspections_subtotal
            patients_total += patients_in_day

            inspections_row += 1

            # summary_sheetの最初のデータの日付まではinspections_sheet単体でのデータ検証を行う
            summary_first_date = return_date(self.summary_sheet.cell(row=main_summary_first_row, column=1).value)
            if date < summary_first_date:
                continue

            summary_inspections = self.summary_sheet.cell(row=main_summary_first_row + count, column=3).value
            summary_patients = self.summary_sheet.cell(row=main_summary_first_row + count, column=4).value
            if inspections_total != summary_inspections:
                add_warning_message(
                    f"{month_and_day(date)}の検査件数に間違いがある可能性があります。" +
                    f"累計が合いません(差分:{summary_inspections - inspections_total})",
                    "summary"
                )
            if patients_total != summary_patients:
                option_file = "summary"
                only_option_file = True

                for warning in patients_warnings:
                    if month_and_day(date) in warning:
                        option_file = ""
                if option_file:
                    add_warning_message(
                        f"{month_and_day(date)}の陽性件数に間違いがある可能性があります。" +
                        f"累計が合いません(差分:{summary_patients - patients_total})",
                        option_file,
                        only_option_file
                    )
            count += 1
        return inspections_warning

    def check_summary_sheet(self) -> List:
        summary_warning = []
        summary_row = main_summary_first_row

        def add_warning_message(message: str, option_file: str = ""):
            summary_warning.append(
                {
                    "message": message,
                    "file": "summary" + (f", {option_file}" if option_file else ""),
                    "fixed": False
                }
            )

        while True:
            date = return_date(self.summary_sheet.cell(row=summary_row, column=1).value)
            if date is None:
                break

            (
                confirmed_cases,
                hospitalized,
                mild,
                severe,
                home_recuperation,
                death,
                discharged
            ) = self.get_summary_values(summary_row)

            # 入院患者数の検証
            if hospitalized != mild + severe:
                add_warning_message(
                    f"{month_and_day(date)}時点の入院患者数に間違いがある可能性があります。" +
                    f"中等症者と重症者の合計と入院患者数が合いません(差分:{hospitalized - (mild + severe)})"
                )
            # 陽性者数の検証
            if confirmed_cases != hospitalized + home_recuperation + death + discharged:
                add_warning_message(
                    f"{month_and_day(date)}時点の陽性者数累計に間違いがある可能性があります。" +
                    "陽性者数累計とその他(入院患者数、自宅療養者数、死者数、退院者数)の合計が合いません" +
                    f"(差分:{confirmed_cases - (hospitalized + home_recuperation + death + discharged)})"
                )
            summary_row += 1

        return summary_warning

    def get_summary_values(self, row) -> List:
        values = []
        for i in range(4, 11):
            value = self.summary_sheet.cell(row=row, column=i).value
            values.append(value)
        return values

    def get_inspections(self) -> None:
        # 検査データの行数の取得
        while self.inspections_sheet:
            self.inspections_count += 1
            value = self.inspections_sheet.cell(row=self.inspections_count, column=1).value
            if not value:
                break

    def slack_notify(self, message: str) -> None:
        requests.post(self.slack_webhook, data=dumps({
            'text': message,
            'username': 'COVID-19 Open Data Validator'
        }))


if __name__ == '__main__':
    print_log("main", "Downloading open data...")
    # データファイルの取得
    # DataManagerだけでなく、DataValidatorでも使用するのでクラスの外に出している
    try:
        patients = get_file("/kk03/corona_kanjyajyokyo.html", True).worksheets[0]
    except AssertionError:
        patients = get_file("/kk03/corona_hasseijyokyo.html", True).worksheets[0]
    inspections = get_file("/kf16/coronavirus_data.html", True, 1).worksheets[0]
    summary = get_file("/kf16/coronavirus_data.html", True).worksheets[0]
    print_log("main", "Complete download of open data.")
    print_log("main", "Init DataManager")
    data_manager = DataManager(patients, inspections, summary)
    changed_flag = data_manager.dump_and_check_all_data()
    # 変更が検知されればlast_update.jsonを生成する
    if changed_flag:
        last_update = {
            "last_update": datetime.now(jst).strftime("%Y-%m-%dT%H:%M:00+09:00")
        }
    else:
        last_update = requests_now_data_json("last_update.json")
    print_log("main", "Make last_update.json...")
    dumps_json("last_update.json", last_update)
    print_log("main", "Make files complete!")
    print_log("main", "Start open data validation.")
    print_log("main", "Init DataValidator")
    data_validator = DataValidator(patients, inspections, summary)
    print_log("main", data_validator.check_all_data())
